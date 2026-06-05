"""分类图集批量导入 — Excel URL + 文件夹 + 去重"""
import io
import hashlib
import logging
from dataclasses import dataclass
from PIL import Image
import requests

logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    filename: str
    status: str  # 'ok' | 'duplicate' | 'error'
    note: str = ''
    data: bytes | None = None


def compute_image_hash(image_data: bytes) -> str:
    return hashlib.sha256(image_data).hexdigest()


def parse_excel_urls(file_data: bytes) -> list[str]:
    """从 Excel 中提取图片链接"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True)
    urls = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip()
                    if val.startswith('http://') or val.startswith('https://'):
                        lower_val = val.lower()
                        if any(ext in lower_val for ext in
                               ['.jpg', '.jpeg', '.png', '.webp', '.gif', '~tplv-', 'image', 'img', 'photo']):
                            urls.append(val)
    wb.close()
    return urls


def download_image(url: str, timeout: int = 30) -> bytes | None:
    """从 URL 下载图片"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        }
        if 'tokopedia-static' in url:
            headers['Referer'] = 'https://www.tokopedia.com/'
        elif 'shopee' in url.lower():
            headers['Referer'] = 'https://shopee.co.id/'
        elif 'lazada' in url.lower():
            headers['Referer'] = 'https://www.lazada.co.id/'

        resp = requests.get(url, headers=headers, timeout=timeout)
        resp.raise_for_status()
        Image.open(io.BytesIO(resp.content)).verify()
        return resp.content
    except Exception as e:
        logger.warning(f'Failed to download {url}: {e}')
        return None


def batch_collect_images(
    files: list | None = None,
    excel_file: bytes | None = None,
) -> tuple[list[ImportResult], list[bytes]]:
    """收集所有图片，去重后返回唯一图片数据列表

    Returns:
        (results, unique_images): 每条记录的结果 + 去重后的图片二进制数据列表
    """
    results = []
    seen_hashes = set()
    unique_images = []

    # 1. 直接上传的文件
    if files:
        for f in files:
            if f.size > 0:
                data = f.read()
                f.seek(0)
                h = compute_image_hash(data)
                if h in seen_hashes:
                    results.append(ImportResult(filename=f.name, status='duplicate', note='重复图片'))
                else:
                    seen_hashes.add(h)
                    results.append(ImportResult(filename=f.name, status='ok', data=data))
                    unique_images.append(data)

    # 2. Excel 中的 URL
    if excel_file:
        urls = parse_excel_urls(excel_file)
        for url in urls:
            filename = url.split('/')[-1].split('?')[0] or 'image.jpg'
            img_data = download_image(url)
            if img_data is None:
                results.append(ImportResult(filename=filename, status='error', note='下载失败'))
                continue
            h = compute_image_hash(img_data)
            if h in seen_hashes:
                results.append(ImportResult(filename=filename, status='duplicate', note='重复图片'))
            else:
                seen_hashes.add(h)
                results.append(ImportResult(filename=filename, status='ok', data=img_data))
                unique_images.append(img_data)

    return results, unique_images
