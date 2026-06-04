"""批量导入印花 — Excel链接 + 文件夹 + 查重 + 自动识别"""
import io
import hashlib
import logging
from dataclasses import dataclass, field
from pathlib import Path
from PIL import Image
import requests

logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    """单张图片导入结果"""
    filename: str
    status: str  # 'new' | 'duplicate' | 'error'
    source_type: str = ''
    note: str = ''
    hash: str = ''


def compute_image_hash(image_data: bytes) -> str:
    """计算图片 SHA256 哈希用于查重"""
    return hashlib.sha256(image_data).hexdigest()


def detect_source_type(image: Image.Image) -> str:
    """自动检测印花图片来源类型

    规则：
    - 透明背景（Alpha通道大量透明像素）→ clean_print
    - 含人物肤色（简单启发式）→ model_photo
    - 其他 → product_photo
    """
    # 确保是 RGBA
    if image.mode != 'RGBA':
        image = image.convert('RGBA')

    width, height = image.size
    pixels = image.load()

    # 统计透明像素
    transparent_count = 0
    total_pixels = width * height

    # 采样检测（大图采样加速）
    sample_step = max(1, min(width, height) // 100)

    for y in range(0, height, sample_step):
        for x in range(0, width, sample_step):
            r, g, b, a = pixels[x, y]
            if a < 128:
                transparent_count += 1

    transparent_ratio = transparent_count / max(1, (total_pixels / (sample_step * sample_step)))

    # 如果超过 40% 像素是透明的，判定为干净印花（提高阈值避免误判）
    if transparent_ratio > 0.40:
        return 'clean_print'

    # 检测肤色（简单启发：大面积的中间色调橙/粉色区域）
    skin_count = 0
    for y in range(0, height, sample_step * 2):
        for x in range(0, width, sample_step * 2):
            r, g, b, a = pixels[x, y]
            if a > 200 and 80 < r < 255 and 40 < g < 220 and 20 < b < 180:
                if r > g > b and r - b > 15:  # 肤色特征：R > G > B
                    skin_count += 1

    skin_ratio = skin_count / max(1, total_pixels / (sample_step * sample_step * 4))

    if skin_ratio > 0.05:
        return 'model_photo'

    return 'product_photo'


def parse_excel_urls(file_data: bytes) -> list[str]:
    """从 Excel 文件中提取图片链接

    遍历所有单元格，提取以 http/https 开头的 URL
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True)
    urls = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip()
                    if val.startswith('http://') or val.startswith('https://'):
                        # 判断是否为图片链接（兼容各种CDN链接格式）
                        lower_val = val.lower()
                        is_image = any([
                            ext in lower_val for ext in
                            ['.jpg', '.jpeg', '.png', '.webp', '.gif', '~tplv-', 'image', 'img', 'photo', 'picture']
                        ])
                        if is_image:
                            urls.append(val)

    wb.close()
    return urls


def download_image(url: str, timeout: int = 30) -> bytes | None:
    """从 URL 下载图片"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        }
        # Tokopedia CDN 需要 Referer
        if 'tokopedia-static' in url:
            headers['Referer'] = 'https://www.tokopedia.com/'
        elif 'shopee' in url.lower():
            headers['Referer'] = 'https://shopee.co.id/'
        elif 'lazada' in url.lower():
            headers['Referer'] = 'https://www.lazada.co.id/'

        resp = requests.get(url, headers=headers, timeout=timeout)
        resp.raise_for_status()

        # 验证是否为有效图片
        Image.open(io.BytesIO(resp.content)).verify()
        return resp.content
    except Exception as e:
        logger.warning(f'Failed to download image from {url}: {e}')
        return None


def batch_import_patterns(
    files: list | None = None,
    excel_file: bytes | None = None,
    uploaded_by=None,
    existing_hashes: set | None = None,
) -> list[ImportResult]:
    """批量导入印花

    Args:
        files: 上传的图片文件列表（Django UploadedFile）
        excel_file: Excel 文件二进制内容
        uploaded_by: 上传用户
        existing_hashes: 已有图片的哈希集合（用于查重）

    Returns:
        导入结果列表
    """
    from apps.patterns.models import Pattern
    from django.core.files.base import ContentFile

    if existing_hashes is None:
        existing_hashes = set(
            Pattern.objects.filter(is_deleted=False)
            .exclude(image_hash='')
            .values_list('image_hash', flat=True)
        )

    results = []
    seen_hashes = set(existing_hashes)  # 本次导入的内部查重

    # 收集所有待导入的图片
    images_to_import = []  # (filename, image_bytes)

    # 1. 直接上传的文件
    if files:
        for f in files:
            if f.size > 0:
                images_to_import.append((f.name, f.read()))

    # 2. Excel 中的 URL
    if excel_file:
        urls = parse_excel_urls(excel_file)
        for url in urls:
            img_data = download_image(url)
            if img_data:
                filename = url.split('/')[-1].split('?')[0] or 'image.jpg'
                images_to_import.append((filename, img_data))

    # 3. 逐张处理
    for filename, img_data in images_to_import:
        try:
            img_hash = compute_image_hash(img_data)

            # 查重
            if img_hash in seen_hashes:
                results.append(ImportResult(
                    filename=filename, status='duplicate',
                    hash=img_hash,
                    note='重复图片（与已有印花哈希相同）'
                ))
                continue

            seen_hashes.add(img_hash)

            # 分析图片
            img = Image.open(io.BytesIO(img_data))
            source_type = detect_source_type(img)
            width, height = img.size

            # 创建 Pattern
            pattern = Pattern(
                uploaded_by=uploaded_by,
                source_type=source_type,
                image_hash=img_hash,
                note=f'{filename} ({width}x{height})',
            )
            pattern.image.save(filename, ContentFile(img_data), save=False)
            pattern.save()

            results.append(ImportResult(
                filename=filename, status='new',
                source_type=source_type, hash=img_hash,
                note=f'{width}x{height}'
            ))

        except Exception as e:
            logger.error(f'Failed to import {filename}: {e}')
            results.append(ImportResult(
                filename=filename, status='error',
                note=str(e)
            ))

    return results
